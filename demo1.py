import requests
import json
import openpyxl
from collections import Counter

if __name__ == '__main__':
    # 设置请求头信息
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3'}

    # 视频的AV号
    aid = 'BV1Sk4y1471G'

    # 获取评论的API链接
    url = f'https://api.bilibili.com/x/v2/reply?jsonp=jsonp&pn=1&type=1&oid={aid}&sort=2'

    # 发送请求
    response = requests.get(url, headers=headers)
    json_data = json.loads(response.text)

    # 获取评论总数
    comment_count = json_data['data']['page']['acount']

    # 获取评论信息
    comments = []
    # for i in range((comment_count - 1) // 20 + 1):
    for i in range((50 - 1) // 20 + 1):
        url = f'https://api.bilibili.com/x/v2/reply?jsonp=jsonp&pn={i + 1}&type=1&oid={aid}&sort=2'
        response = requests.get(url, headers=headers)
        json_data = json.loads(response.text)
        for comment in json_data['data']['replies']:
            comments.append(comment['content']['message'])

    # 分词
    words = []
    for comment in comments:
        words += comment.split()

    # 统计高频词汇
    word_counts = Counter(words)
    word_list = [(count, word) for word, count in word_counts.items()]
    word_list.sort(reverse=True)

    # 写入Excel文件
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '评论'
    sheet.cell(1, 1, '评论')
    for i, comment in enumerate(comments):
        sheet.cell(i + 2, 1, comment)

    # 将高频词汇写入Excel文件
    sheet_name = '高频词汇'
    sheet = workbook.create_sheet(sheet_name)
    sheet.cell(1, 1, '出现次数')
    sheet.cell(1, 2, '词汇')
    for i, (count, word) in enumerate(word_list):
        sheet.cell(i + 2, 1, count)
        sheet.cell(i + 2, 2, word)

    # 保存Excel文件
    workbook.save(f'{aid}.xlsx')
